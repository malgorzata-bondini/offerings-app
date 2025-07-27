import datetime as dt
import re
import time
import warnings
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from dataclasses import dataclass
from typing import List, Dict, Optional, Any
import streamlit as st  # Add this import

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

need_cols = [
    "Name (Child Service Offering lvl 1)", "Parent Offering",
    "Service Offerings | Depend On (Application Service)", "Service Commitments",
    "Delivery Manager", "Subscribed by Location", "Phase", "Status",
    "Life Cycle Stage", "Life Cycle Status", "Support group", "Managed by Group",
    "Subscribed by Company", "Business Criticality",
    "Record view", "Approval required", "Approval group"  # Removed "Visibility group"
]

discard_lc = {"retired", "retiring", "end of life", "end of support"}

def ensure_incident_naming(name):
    """
    Ensure that if any keyword is 'incident', then 'solving' is right after 'incident' and then app name
    This function reorganizes the name to ensure proper order: incident solving [app] [other parts]
    """
    # First, check if we already have "incident solving" in the correct order
    if re.search(r'\bincident\s+solving\b', name, re.IGNORECASE):
        return name
    
    # Split the name into parts
    parts = name.split()
    new_parts = []
    i = 0
    
    while i < len(parts):
        if parts[i].lower() == "incident":
            # Add "incident solving"
            new_parts.append(parts[i])
            new_parts.append("solving")
            
            # Skip if the next word was already "solving"
            if i + 1 < len(parts) and parts[i + 1].lower() == "solving":
                i += 1
        elif parts[i].lower() == "solving":
            # Skip standalone "solving" as we'll add it after "incident"
            pass
        else:
            new_parts.append(parts[i])
        i += 1
    
    # Now we need to ensure app name comes AFTER "incident solving", not between
    # Find if there's an app name that got placed between incident and solving
    final_parts = []
    i = 0
    
    while i < len(new_parts):
        if i > 0 and new_parts[i-1].lower() == "incident" and new_parts[i].lower() == "solving":
            # We found "incident solving" - good
            final_parts.append(new_parts[i])
        elif new_parts[i].lower() == "incident":
            # Check if there's something between incident and solving
            j = i + 1
            app_parts = []
            
            # Collect everything until we find "solving"
            while j < len(new_parts) and new_parts[j].lower() != "solving":
                app_parts.append(new_parts[j])
                j += 1
            
            # Add incident solving first
            final_parts.append(new_parts[i])
            if j < len(new_parts) and new_parts[j].lower() == "solving":
                final_parts.append("solving")
                # Then add the app parts that were in between
                final_parts.extend(app_parts)
                i = j
            else:
                # No solving found, just add solving after incident
                final_parts.append("solving")
                final_parts.extend(app_parts)
                i = j - 1
        else:
            final_parts.append(new_parts[i])
        i += 1
    
    return " ".join(final_parts)

def extract_parent_info(parent_offering):
    """Extract the content between [Parent ...] from parent offering"""
    match = re.search(r'\[Parent\s+(.*?)\]', str(parent_offering), re.I)
    if match:
        return match.group(1).strip()
    return ""

def extract_catalog_name(parent_offering):
    """Extract the catalog name after the brackets"""
    parts = str(parent_offering).split(']', 1)
    if len(parts) > 1:
        return parts[1].strip()
    return ""

def get_division_and_country(parent_content, country, delivering_tag):
    """Get division and country with special handling for MD, UA, RO, and TR"""
    # Special handling for UA, MD, RO, and TR - always use DS
    if country in ["UA", "MD", "RO", "TR"]:
        return "DS", country
    
    # Extract parts from parent content
    parts = parent_content.split()
    division = ""
    
    for part in parts:
        if part in ["HS", "DS"]:
            division = part
            break
    
    # If no division found in parent and we have delivering_tag, use it
    if not division and delivering_tag:
        delivering_parts = delivering_tag.split()
        if delivering_parts and delivering_parts[0] in ["HS", "DS"]:
            division = delivering_parts[0]
    
    # Default to HS if still no division
    if not division:
        division = "HS"
    
    return division, country

def build_lvl2_name(parent_offering, sr_or_im, app, schedule_suffix, service_type_lvl2):
    """Build name for Lvl2 entries - SR/IM is added to both Parent Offering parsing and final name"""
    parent_content = extract_parent_info(parent_offering)
    catalog_name = extract_catalog_name(parent_offering)
    
    # Check if SR/IM already exists in parent content
    parts = parent_content.split()
    has_sr_im = "SR" in parts or "IM" in parts
    
    # If SR/IM not already there, insert it after Parent
    if not has_sr_im:
        # Insert SR/IM as the first element after "Parent" was removed
        parts.insert(0, sr_or_im)
        parent_content = " ".join(parts)
    
    # Now extract parts for building
    parts = parent_content.split()
    country = ""
    division = ""
    dept = ""  # IT, HR, etc.
    sr_im_pos = ""
    
    for part in parts:
        if part in ["SR", "IM"]:
            sr_im_pos = part
        elif len(part) == 2 and part.isupper() and part not in ["HS", "DS", "IT", "HR"]:
            country = part
        elif part in ["HS", "DS"]:
            division = part
        elif part in ["IT", "HR", "Medical", "Business Services"]:
            dept = part
    
    # Build the prefix - use the SR/IM from parent or the one provided
    prefix_parts = [sr_im_pos if sr_im_pos else sr_or_im]
    
    # Special handling for UA, MD, RO, and TR - always use DS
    if country in ["UA", "MD", "RO", "TR"]:
        prefix_parts.append("DS")
    elif division:
        prefix_parts.append(division)
    else:
        prefix_parts.append("HS")  # Default
    
    if country:
        prefix_parts.append(country)
    
    if dept:
        prefix_parts.append(dept)
    
    # Build the name
    name_prefix = f"[{' '.join(prefix_parts)}]"
    
    # Extract the core part of catalog name (e.g., "Software incident solving")
    name_parts = [name_prefix, catalog_name]
    
    # Add app if provided
    if app:
        name_parts.append(app)
    
    # Check if name contains Microsoft - if so, don't add Prod
    name_so_far = " ".join(name_parts).lower()
    if "microsoft" not in name_so_far:
        name_parts.append("Prod")
    
    # Add service type if provided (e.g., "Application issue")
    if service_type_lvl2:
        name_parts.append(service_type_lvl2)
    
    # Add schedule
    name_parts.append(schedule_suffix)
    
    # Join and ensure incident naming
    final_name = " ".join(name_parts)
    return ensure_incident_naming(final_name)

def build_corp_it_name(parent_offering, sr_or_im, app, schedule_suffix, receiver, delivering_tag):
    """Build name for CORP IT offerings"""
    parent_content = extract_parent_info(parent_offering)
    catalog_name = extract_catalog_name(parent_offering)
    
    # Extract parts from parent content
    parts = parent_content.split()
    country = ""
    topic = ""
    
    for part in parts:
        if len(part) == 2 and part.isupper() and part not in ["HS", "DS"]:
            country = part
        elif part not in ["HS", "DS", "Parent", "RecP"] and not (len(part) == 2 and part.isupper()):
            topic = part
            break
    
    # Build CORP IT name - always ends with IT
    prefix_parts = [sr_or_im]
    
    # Get division and country with special handling for MD/UA/RO/TR
    division, country_code = get_division_and_country(parent_content, country, delivering_tag)
    
    # Add delivering tag parts (who delivers the service - from user input)
    if delivering_tag:
        delivering_parts = delivering_tag.split()
        # For MD/UA/RO/TR, override with DS
        if country in ["UA", "MD", "RO", "TR"]:
            prefix_parts.extend(["DS", country])
        else:
            prefix_parts.extend(delivering_parts)
    else:
        prefix_parts.extend([division, country])
    
    prefix_parts.append("CORP")
    
    # Add receiver parts (who receives the service - DS DE or HS DE for Germany)
    if receiver:
        prefix_parts.append(receiver)
    else:
        prefix_parts.extend([division, country])
    
    prefix_parts.append("IT")
    
    # Build the name
    name_prefix = f"[{' '.join(prefix_parts)}]"
    
    # Use topic from parent or "Software" as default
    if not topic:
        topic = "Software"
    
    name_parts = [name_prefix, topic, catalog_name.lower()]
    
    # Add app if provided
    if app:
        name_parts.append(app)
    
    # Add "solving" for IM
    if sr_or_im == "IM":
        name_parts.append("solving")
    
    # name_parts.append("Prod")
    name_parts.append(schedule_suffix)
    
    # Join and ensure incident naming
    final_name = " ".join(name_parts)
    return ensure_incident_naming(final_name)

def build_corp_dedicated_name(parent_offering, sr_or_im, app, schedule_suffix, receiver, delivering_tag):
    """Build name for CORP Dedicated Services offerings"""
    parent_content = extract_parent_info(parent_offering)
    catalog_name = extract_catalog_name(parent_offering)
    
    # Extract parts from parent content
    parts = parent_content.split()
    country = ""
    topic = ""
    
    for part in parts:
        if len(part) == 2 and part.isupper() and part not in ["HS", "DS"]:
            country = part
        elif part not in ["HS", "DS", "Parent", "RecP"] and not (len(part) == 2 and part.isupper()):
            topic = part
            break
    
    # Build CORP Dedicated Services name
    prefix_parts = [sr_or_im]
    
    # Get division and country with special handling for MD/UA/RO/TR
    division, country_code = get_division_and_country(parent_content, country, delivering_tag)
    
    # Add delivering tag parts (who delivers the service - from user input)
    if delivering_tag:
        delivering_parts = delivering_tag.split()
        # For MD/UA/RO/TR, override with DS
        if country in ["UA", "MD", "RO", "TR"]:
            prefix_parts.extend(["DS", country])
        else:
            prefix_parts.extend(delivering_parts)
    else:
        prefix_parts.extend([division, country])
    
    prefix_parts.append("CORP")
    
    # Add receiver (e.g., HS DE)
    prefix_parts.append(receiver)
    prefix_parts.append("Dedicated Services")
    
    # Build the name
    name_prefix = f"[{' '.join(prefix_parts)}]"
    
    name_parts = [name_prefix, catalog_name]
    
    # Add app if provided
    if app:
        name_parts.append(app)
    
    # Add "solving" for IM
    if sr_or_im == "IM":
        name_parts.append("solving")
    
    name_parts.append("Prod")
    name_parts.append(schedule_suffix)
    
    # Join and ensure incident naming
    final_name = " ".join(name_parts)
    return ensure_incident_naming(final_name)

def build_recp_name(parent_offering, sr_or_im, app, schedule_suffix, receiver, delivering_tag):
    """Build name for RecP offerings"""
    parent_content = extract_parent_info(parent_offering)
    catalog_name = extract_catalog_name(parent_offering)
    
    # Extract parts from parent content
    parts = parent_content.split()
    country = ""
    topic = ""
    
    for part in parts:
        if len(part) == 2 and part.isupper() and part not in ["HS", "DS"]:
            country = part
        elif part not in ["HS", "DS", "Parent", "RecP"] and not (len(part) == 2 and part.isupper()):
            topic = part
            break
    
    # Build RecP name - always ends with IT
    prefix_parts = [sr_or_im]
    
    # Get division and country with special handling for MD/UA/RO/TR
    division, country_code = get_division_and_country(parent_content, country, delivering_tag)
    
    # Extract division and country from parent
    parent_division = ""
    for part in parts:
        if part in ["HS", "DS"]:
            parent_division = part
            break
    
    # For MD/UA/RO/TR, always use DS
    if country in ["UA", "MD", "RO", "TR"]:
        prefix_parts.extend(["DS", country])
    else:
        if parent_division:
            prefix_parts.append(parent_division)
        if country:
            prefix_parts.append(country)
    
    prefix_parts.append("CORP")
    
    # Add delivering tag parts
    if delivering_tag:
        delivering_parts = delivering_tag.split()
        # For MD/UA/RO/TR, override with DS
        if country in ["UA", "MD", "RO", "TR"]:
            prefix_parts.extend(["DS", country])
        else:
            prefix_parts.extend(delivering_parts)
    else:
        # For MD/UA/RO/TR, use DS
        if country in ["UA", "MD", "RO", "TR"]:
            prefix_parts.extend(["DS", country])
        else:
            prefix_parts.extend([division, country])
    
    prefix_parts.append("IT")
    
    # Build the name with topic from parent
    name_prefix = f"[{' '.join(prefix_parts)}]"
    
    # For RecP, use topic (e.g., "Software") + catalog name in lowercase
    name_parts = [name_prefix]
    
    if topic:
        name_parts.append(topic)
    
    name_parts.append(catalog_name.lower())
    
    # Add app if provided
    if app:
        name_parts.append(app)
    
    # Add "solving" for IM
    if sr_or_im == "IM":
        name_parts.append("solving")
    
    name_parts.append("Prod")
    name_parts.append(schedule_suffix)
    
    # Join and ensure incident naming
    final_name = " ".join(name_parts)
    return ensure_incident_naming(final_name)

def build_standard_name(parent_offering, sr_or_im, app, schedule_suffix, special_dept=None, receiver=None):
    """Build standard name when not CORP"""
    parent_content = extract_parent_info(parent_offering)
    catalog_name = extract_catalog_name(parent_offering)
    
    # Apply pluralization to catalog name for hardware items
    catalog_name_plural = get_plural_form(catalog_name)
    
    # Extract country from parent content
    parts = parent_content.split()
    country = ""
    for part in parts:
        if len(part) == 2 and part.isupper() and part not in ["HS", "DS", "IT", "HR"]:
            country = part
            break
    
    # Check if catalog name, parent offering, or parent content contains keywords that exclude "Prod"
    no_prod_keywords = ["hardware", "mailbox", "network", "mobile", "security"]
    parent_lower = parent_offering.lower()
    catalog_lower = catalog_name.lower()
    parent_content_lower = parent_content.lower()
    exclude_prod = any(keyword in parent_lower or keyword in catalog_lower or keyword in parent_content_lower 
                      for keyword in no_prod_keywords)
    
    if special_dept == "Medical":
        # Extract division and country from parent content
        parts = parent_content.split()
        division = ""
        country = ""
        topic_parts = []
        
        for i, part in enumerate(parts):
            if part in ["HS", "DS"]:
                division = part
            elif len(part) == 2 and part.isupper() and part not in ["IT", "HR"]:
                country = part
            elif part not in ["HS", "DS"] and not (len(part) == 2 and part.isupper()):
                topic_parts.append(part)
        
        topic = " ".join(topic_parts) if topic_parts else "Software"
        
        # Build Medical name - NO PROD
        prefix_parts = [sr_or_im]
        
        # Special handling for UA, MD, RO, and TR - always use DS
        if country in ["UA", "MD", "RO", "TR"]:
            prefix_parts.extend(["DS", country])
        else:
            if division:
                prefix_parts.append(division)
            if country:
                prefix_parts.append(country)
        
        prefix_parts.append("Medical")
        
        # Use topic from parent and lowercase catalog name
        final_name = f"[{' '.join(prefix_parts)}] {topic} {catalog_name.lower()} {schedule_suffix}"
        return ensure_incident_naming(final_name)
    
    elif special_dept == "DAK":
        # Replace DAK with Business Services - NO PROD
        parts = parent_content.split()
        division = ""
        country = ""
        
        for part in parts:
            if part in ["HS", "DS"]:
                division = part
            elif len(part) == 2 and part.isupper() and part not in ["IT", "HR", "DAK"]:
                country = part
        
        prefix_parts = [sr_or_im]
        
        # Special handling for UA, MD, RO, and TR - always use DS
        if country in ["UA", "MD", "RO", "TR"]:
            prefix_parts.extend(["DS", country])
        else:
            if division:
                prefix_parts.append(division)
            if country:
                prefix_parts.append(country)
        
        prefix_parts.append("Business Services")
        
        # Add app only if provided - NO PROD
        if app:
            final_name = f"[{' '.join(prefix_parts)}] {catalog_name} {app} {schedule_suffix}"
        else:
            final_name = f"[{' '.join(prefix_parts)}] {catalog_name} {schedule_suffix}"
        return ensure_incident_naming(final_name)
    
    elif special_dept == "HR":
        # HR - NO PROD
        parts = parent_content.split()
        division = ""
        country = ""
        for part in parts:
            if part in ["HS", "DS"]:
                division = part
            elif len(part) == 2 and part.isupper() and part not in ["IT", "HR"]:
                country = part
        
        prefix_parts = [sr_or_im]
        
        # Special handling for UA, MD, RO, and TR - always use DS
        if country in ["UA", "MD", "RO", "TR"]:
            prefix_parts.extend(["DS", country])
        else:
            if division:
                prefix_parts.append(division)
            if country:
                prefix_parts.append(country)
        
        prefix_parts.append("HR")
        
        # Extract the topic from parent content
        topic_parts = []
        for part in parts:
            if part not in ["HS", "DS"] and not (len(part) == 2 and part.isupper()):
                topic_parts.append(part)
        
        topic = " ".join(topic_parts) if topic_parts else "Software"
        
        # Add app if provided - NO PROD
        if app:
            final_name = f"[{' '.join(prefix_parts)}] {topic} {catalog_name.lower()} {app} {schedule_suffix}"
        else:
            final_name = f"[{' '.join(prefix_parts)}] {topic} {catalog_name.lower()} {schedule_suffix}"
        return ensure_incident_naming(final_name)
    
    elif special_dept == "IT":
        # IT - special handling
        parts = parent_content.split()
        division = ""
        country = ""
        topic = ""
        
        # Find division, country and topic
        topic_parts = []
        for i, part in enumerate(parts):
            if part in ["HS", "DS"]:
                division = part
            elif len(part) == 2 and part.isupper() and part not in ["IT", "HR"]:
                country = part
            else:
                # Collect all remaining words as topic (e.g., "Security & Privacy", "Hardware", etc.)
                if part not in ["HS", "DS", "RecP"] and not (len(part) == 2 and part.isupper()):
                    topic_parts.append(part)
        
        # Join all topic parts to get the full topic phrase
        topic = " ".join(topic_parts) if topic_parts else ""
        
        # If no topic found, use the first significant word from catalog name
        if not topic:
            catalog_words = catalog_name.split()
            for word in catalog_words:
                if word.lower() not in ["the", "a", "an", "and", "or", "for", "of", "in", "on", "to"]:
                    topic = word
                    break
        
        prefix_parts = [sr_or_im]
        
        # For DE with receiver, extract division from receiver (e.g., "HS DE" -> "HS")
        if country == "DE" and receiver:
            recv_division = receiver.split()[0]  # Extract HS or DS
            prefix_parts.append(recv_division)
        # Special handling for UA, MD, RO, and TR - always use DS
        elif country in ["UA", "MD", "RO", "TR"]:
            prefix_parts.append("DS")
        elif division:
            prefix_parts.append(division)
        else:
            # Default to HS if no division found
            prefix_parts.append("HS")
        
        if country:
            prefix_parts.append(country)
        prefix_parts.append("IT")
        
        # For IT, build the name with topic BEFORE the brackets
        # Format: [SR DS MD IT] Hardware configuration laptop Mon-Fri 8-16
        name_parts = []
        
        if topic:
            # Topic goes BEFORE the brackets
            name_parts.append(f"[{' '.join(prefix_parts)}] {topic}")
        else:
            name_parts.append(f"[{' '.join(prefix_parts)}]")
        
        name_parts.append(catalog_name.lower())
        
        # Add app if provided
        if app:
            # Build the current name string to check for "hardware"
            current_name_str = " ".join(name_parts)
            # Check if "hardware" is in the current name (case insensitive)
            if "hardware" in current_name_str.lower():
                # Use lowercase for hardware, but keep UPS uppercase
                if app.upper() == "UPS":
                    name_parts.append("UPS")  # Keep UPS uppercase
                else:
                    name_parts.append(app.lower())  # Use lowercase for other hardware
            else:
                name_parts.append(app)  # Keep original case for non-hardware
        
        # Add "solving" for IM
        if sr_or_im == "IM":
            name_parts.append("solving")
        
        # Check if topic contains any no-prod keywords
        topic_lower = topic.lower() if topic else ""
        topic_exclude_prod = any(keyword in topic_lower for keyword in no_prod_keywords)
        
        # Only add Prod if no hardware/mailbox/network/mobile/security keywords in any source
        if not exclude_prod and not topic_exclude_prod:
            name_parts.append("Prod")
        
        name_parts.append(schedule_suffix)
        
        # Join and ensure incident naming
        final_name = " ".join(name_parts)
        return ensure_incident_naming(final_name)
    
    else:
        # Standard case - replace Parent with SR/IM and add IT for RecP entries
        parts = parent_content.split()
        division = ""
        country_code = ""
        dept = ""
        other_parts = []
        
        # Parse parent content to extract components
        for part in parts:
            if part in ["HS", "DS"]:
                division = part
            elif len(part) == 2 and part.isupper() and part not in ["HS", "DS", "IT", "HR"]:
                country_code = part
            elif part in ["IT", "HR", "Medical", "Business Services"]:
                dept = part
            else:
                other_parts.append(part)
        
        # Build the new name components
        name_parts = [sr_or_im]
        
        # Special handling for UA, MD, RO, and TR - always use DS
        if country in ["UA", "MD", "RO", "TR"]:
            name_parts.append("DS")
        elif division:
            name_parts.append(division)
        else:
            # Default to HS if no division found
            name_parts.append("HS")
        
        if country_code:
            name_parts.append(country_code)
        
        # Add other parts (like RecP, Software, etc.)
        name_parts.extend(other_parts)
        
        # For RecP entries, add IT department if not already present
        if "RecP" in other_parts and not dept:
            name_parts.append("IT")
        elif dept:
            name_parts.append(dept)
        
        # Build the final name
        prefix = f"[{' '.join(name_parts)}]"
        
        # Add catalog name
        final_parts = [prefix, catalog_name]
        
        # Check if we should add Prod
        catalog_lower = catalog_name.lower()
        exclude_prod = any(keyword in catalog_lower for keyword in ["hardware", "mailbox", "network", "mobile", "security"])
        
        # Add app if provided
        if app:
            # Build the current name string to check for "hardware"
            current_name_str = " ".join(final_parts)
            # Check if "hardware" is in the current name (case insensitive)
            if "hardware" in current_name_str.lower():
                # Use lowercase for hardware, but keep UPS uppercase
                if app.upper() == "UPS":
                    final_parts.append("UPS")  # Keep UPS uppercase
                else:
                    final_parts.append(app.lower())  # Use lowercase for other hardware
            else:
                final_parts.append(app)  # Keep original case for non-hardware
        
        # Add "solving" for IM
        if sr_or_im == "IM":
            final_parts.append("solving")
        
        if not exclude_prod:
            final_parts.append("Prod")
        
        final_parts.append(schedule_suffix)
        
        final_name = " ".join(final_parts)
        return ensure_incident_naming(final_name)

def build_corp_name(parent_offering, sr_or_im, app, schedule_suffix, receiver, delivering_tag):
    """Build name for CORP offerings"""
    parent_content = extract_parent_info(parent_offering)
    catalog_name = extract_catalog_name(parent_offering)
    
    # Extract parts from parent content
    parts = parent_content.split()
    country = ""
    topic = ""
    
    for part in parts:
        if len(part) == 2 and part.isupper() and part not in ["HS", "DS"]:
            country = part
        elif part not in ["HS", "DS", "Parent", "RecP"]:
            topic = part
    
    # Build CORP name
    prefix_parts = [sr_or_im]
    
    # Get division and country with special handling for MD/UA/RO/TR
    division, country_code = get_division_and_country(parent_content, country, delivering_tag)
    
    # Add delivering tag parts (who delivers the service - from user input)
    if delivering_tag:
        delivering_parts = delivering_tag.split()
        # For MD/UA/RO/TR, override with DS
        if country in ["UA", "MD", "RO", "TR"]:
            prefix_parts.extend(["DS", country])
        else:
            prefix_parts.extend(delivering_parts[:2])  # Take division and country from delivering tag
    else:
        prefix_parts.extend([division, country])
    
    prefix_parts.extend(["CORP", receiver])
    
    # Only add topic if it exists, don't default to IT for CORP
    if topic:
        prefix_parts.append(topic)
    
    if sr_or_im == "SR":
        if app:
            final_name = f"[{' '.join(prefix_parts)}] {catalog_name} {app} Prod {schedule_suffix}"
        else:
            final_name = f"[{' '.join(prefix_parts)}] {catalog_name} Prod {schedule_suffix}"
    else:
        # For IM: always add solving after the app (or after catalog if no app)
        if app:
            final_name = f"[{' '.join(prefix_parts)}] {catalog_name} {app} solving Prod {schedule_suffix}"
        else:
            final_name = f"[{' '.join(prefix_parts)}] {catalog_name} solving Prod {schedule_suffix}"
    
    return ensure_incident_naming(final_name)

def commit_block(cc, schedule_suffix, rsp_duration, rsl_duration, sr_or_im):
    """Create commitment block with OLA for all countries"""
    if sr_or_im == "IM":
        # For IM, no OLA
        lines = [
            f"[{cc}] SLA IM RSP {schedule_suffix} P1-P4 {rsp_duration}",
            f"[{cc}] SLA IM RSL {schedule_suffix} P1-P4 {rsl_duration}"
        ]
    else:
        # For SR, include OLA (but only once)
        lines = [
            f"[{cc}] SLA SR RSP {schedule_suffix} P1-P4 {rsp_duration}",
            f"[{cc}] SLA SR RSL {schedule_suffix} P1-P4 {rsl_duration}",
            f"[{cc}] OLA SR RSL {schedule_suffix} P1-P4 {rsl_duration}"
        ]
    return "\n".join(lines)

def update_commitments(orig, sched, rsp, rsl, sr_or_im, country):
    """Update existing commitments and ensure OLA is present"""
    out = []
    has_ola = False
    country_code = None
    
    for line in str(orig).splitlines():
        line = line.strip()
        if not line:
            continue
            
        if "RSP" in line:
            # Extract country code from line like [PL] SLA SR RSP...
            match = re.search(r'\[(\w+)\]', line)
            if match:
                country_code = match.group(1)
            # Extract P values (P1-P4, P1-P3, etc)
            p_match = re.search(r'(P\d+-P\d+)', line)
            p_values = p_match.group(1) if p_match else "P1-P4"
            # Update schedule and duration
            line = re.sub(r"RSP\s+[^P]+", f"RSP {sched} ", line)
            line = re.sub(r"(P\d+-P\d+)\s+.*$", f"{p_values} {rsp}", line)
        elif "RSL" in line:
            # Extract P values
            p_match = re.search(r'(P\d+-P\d+)', line)
            p_values = p_match.group(1) if p_match else "P1-P4"
            # Update schedule and duration
            line = re.sub(r"RSL\s+[^P]+", f"RSL {sched} ", line)
            line = re.sub(r"(P\d+-P\d+)\s+.*$", f"{p_values} {rsl}", line)
        elif "OLA" in line:
            has_ola = True
            # Extract P values
            p_match = re.search(r'(P\d+-P\d+)', line)
            p_values = p_match.group(1) if p_match else "P1-P4"
            # Update schedule and duration - OLA uses same pattern as RSL
            line = re.sub(r"RSL\s+[^P]+", f"RSL {sched} ", line)
            line = re.sub(r"(P\d+-P\d+)\s+.*$", f"{p_values} {rsl}", line)
        out.append(line)
    
    return "\n".join(out)

def custom_commit_block(cc, sr_or_im, rsp_enabled, rsl_enabled, rsp_schedule, rsl_schedule, 
                       rsp_priority, rsl_priority, rsp_time, rsl_time):
    """Create custom commitment block based on user selections"""
    lines = []
    
    if rsp_enabled and rsp_schedule and rsp_priority and rsp_time:
        lines.append(f"[{cc}] SLA {sr_or_im} RSP {rsp_schedule} {rsp_priority} {rsp_time}")
    
    if rsl_enabled and rsl_schedule and rsl_priority and rsl_time:
        lines.append(f"[{cc}] SLA {sr_or_im} RSL {rsl_schedule} {rsl_priority} {rsl_time}")
        # Add OLA for SR only
        if sr_or_im == "SR":
            lines.append(f"[{cc}] OLA {sr_or_im} RSL {rsl_schedule} {rsl_priority} {rsl_time}")
    
    return "\n".join(lines) if lines else ""

def create_new_parent_row(new_parent_offering, new_parent, country, business_criticality="", approval_required=False, approval_required_value="empty", change_subscribed_location=False, custom_subscribed_location="Global"):
    """Create a new row with the specified parent offering and parent values"""
    # Create a basic row structure with required columns
    new_row = {
        "Name (Child Service Offering lvl 1)": "",  # Will be filled later
        "Parent Offering": new_parent_offering,  # Use the user-provided value
        "Parent": new_parent,  # Use the user-provided value (not hardcoded)
        "Service Offerings | Depend On (Application Service)": "",
        "Service Commitments": "",
        "Delivery Manager": "",
        "Subscribed by Location": "",  # Will be set below
        "Phase": "Catalog",
        "Status": "Operational",
        "Life Cycle Stage": "Operational",
        "Life Cycle Status": "In Use",
        "Support group": "",
        "Managed by Group": "",
        "Subscribed by Company": "",  # Will be set during processing based on receiver and CORP type
        "Business Criticality": business_criticality,
        "Record view": "",  # Will be set based on SR/IM
        "Approval required": "true" if approval_required else "false",  # Always use "true"/"false"
        "Approval group": approval_required_value if approval_required else "empty"  # Use custom value for approval group
        # Removed "Visibility group" line
    }
    
    # Set Subscribed by Location based on user choice
    if change_subscribed_location:
        new_row["Subscribed by Location"] = custom_subscribed_location
    else:
        new_row["Subscribed by Location"] = "Global"
    
    return pd.Series(new_row)

def get_support_group_for_country(country, support_group, support_groups_per_country, division=None):
    """Get the appropriate support group for a given country and division"""
    # For PL, use division-specific support groups
    if country == "PL" and division:
        division_key = f"{division} PL"
        if support_groups_per_country and division_key in support_groups_per_country:
            return support_groups_per_country[division_key]
    
    # For other countries, use the country key directly
    if support_groups_per_country and country in support_groups_per_country:
        return support_groups_per_country[country]
    return support_group

def get_managed_by_group_for_country(country, managed_by_group, managed_by_groups_per_country, 
                                     support_group_for_country, division=None):
    """Get the appropriate managed by group for a given country and division"""
    # For PL, use division-specific managed by groups
    if country == "PL" and division:
        division_key = f"{division} PL"
        if managed_by_groups_per_country and division_key in managed_by_groups_per_country:
            managed_by_value = managed_by_groups_per_country[division_key]
            # If managed_by is empty but support_group is filled, use support_group
            return managed_by_value if managed_by_value else support_group_for_country
    
    # For other countries, use the country key directly
    if managed_by_groups_per_country and country in managed_by_groups_per_country:
        managed_by_value = managed_by_groups_per_country[country]
        # If managed_by is empty but support_group is filled, use support_group
        return managed_by_value if managed_by_value else support_group_for_country
    # Fall back to global managed_by_group, or support_group_for_country if empty
    return managed_by_group if managed_by_group else support_group_for_country

def get_support_groups_list_for_country(country, support_group, support_groups_per_country, 
                                       managed_by_groups_per_country, division=None):
    """Get list of support groups for a country (handles multiple groups for DE)"""
    # Get support groups
    if country == "PL" and division:
        division_key = f"{division} PL"
        country_support_groups = support_groups_per_country.get(division_key, support_group) if support_groups_per_country else support_group
        country_managed_groups = managed_by_groups_per_country.get(division_key, "") if managed_by_groups_per_country else ""
    else:
        country_support_groups = support_groups_per_country.get(country, support_group) if support_groups_per_country else support_group
        country_managed_groups = managed_by_groups_per_country.get(country, "") if managed_by_groups_per_country else ""
    
    # Handle multiple groups (separated by newlines)
    if country_support_groups and '\n' in str(country_support_groups):
        support_list = [sg.strip() for sg in str(country_support_groups).split('\n') if sg.strip()]
        managed_list = []
        if country_managed_groups and '\n' in str(country_managed_groups):
            managed_list = [mg.strip() for mg in str(country_managed_groups).split('\n') if mg.strip()]
        else:
            managed_list = [country_managed_groups.strip()] * len(support_list) if country_managed_groups else support_list
        
        # Ensure managed_list has same length as support_list
        while len(managed_list) < len(support_list):
            managed_list.append(support_list[len(managed_list)])
            
        return list(zip(support_list, managed_list))
    else:
        # Single support group
        single_support = country_support_groups or support_group or ""
        single_managed = country_managed_groups or single_support or ""
        return [(single_support, single_managed)] if single_support else [("", "")]

def get_schedule_suffixes_for_country(country, receiver, schedule_settings_per_country, default_schedule_suffixes):
    """Get the appropriate schedule suffixes for a given country and receiver"""
    # For countries that split into DS/HS (PL), use receiver key
    # For DE, it doesn't split in the same way, so use country directly
    # For CY (now DS-only), RO, TR, use country directly
    if country in ["PL"] and receiver:
        key = receiver  # e.g., "HS PL", "DS PL"
    else:
        key = country  # e.g., "DE", "MD", "UA", "CY", "RO", "TR"
    
    # Check if there are custom schedules for this country/receiver
    if key in schedule_settings_per_country:
        custom_schedules = schedule_settings_per_country[key]
        if isinstance(custom_schedules, str):
            # Split on newlines if it's a string
            return [s.strip() for s in custom_schedules.split('\n') if s.strip()]
        elif isinstance(custom_schedules, list):
            return custom_schedules
    
    # Fallback to default schedule suffixes
    return default_schedule_suffixes

def get_de_company_and_ldap(support_group, receiver, original_row=None):
    """Get the Subscribed by Company and LDAP values for DE based on support group"""
    # Normalize the support group name for comparison (remove extra spaces, normalize case)
    normalized_sg = ' '.join(support_group.strip().split()) if support_group else ""
    
    # Special mappings for DE support groups - using normalized comparison
    if "HS DE IT Service Desk HC" in normalized_sg and receiver == "HS DE":
        return "DE Internal Patients", "CALDOM1.DE [Hospital Calbe]"
    elif "HS DE IT Service Desk - MCC" in normalized_sg and receiver == "HS DE":
        return "DE External Patients", "mednet-de.world [Medicover Clinics]"
    elif ("DS DE IT Service Desk -Labs" in normalized_sg or "DS DE IT Service Desk - Labs" in normalized_sg) and receiver == "DS DE":
        return "DE IFLB Laboratories\nDE IMD Laboratories", "imd-labore.intern [General]"
    else:
        # For other support groups, return the original "Subscribed by Company" value if available
        if original_row is not None and "Subscribed by Company" in original_row.index:
            original_company = str(original_row["Subscribed by Company"]).strip()
            if original_company and original_company not in ["nan", "NaN", "", "None"]:
                return original_company, ""
        # Fallback to support group name if no original value available
        return support_group, ""

PLURAL_MAP = {
    "Laptop": "Laptops",
    "Desktop": "Desktops", 
    "Docking station": "Docking stations",
    "Printer": "Printers",
    "Barcode printer": "Barcode printers",
    "Barcode scanner": "Barcode scanners",
    "Display": "Displays",
    "Deskphone": "Deskphones",
    "Smartphone": "Smartphones",
    "Mouse": "Mouses",  # Note: "Mice" would be grammatically correct, but "Mouses" is common in IT
    "Keyboard": "Keyboards",
    "Headset": "Headsets",
    "Tablet": "Tablets",
    "Audio equipment": "Audio equipment",  # Already plural
    "Video surveillance": "Video surveillance",  # Already plural/uncountable
    "UPS": "UPS",  # Acronym, same in plural
    "External webcam": "External webcams",
    "Projector": "Projectors", 
    "External storage device": "External storage devices",
    "Microphone": "Microphones",
    "Other hardware": "Other hardware",  # Already plural/uncountable
    # Add more as needed
    "Server": "Servers",
    "Router": "Routers",
    "Switch": "Switches",
    "Firewall": "Firewalls",
    "Access point": "Access points",
    "Scanner": "Scanners",
    "Webcam": "Webcams",
    "Camera": "Cameras",
    "Monitor": "Monitors",
    "Speaker": "Speakers",
    "Cable": "Cables",
    "Adapter": "Adapters"
}

def get_plural_form(word):
    """Get plural form of a word if it exists in PLURAL_MAP, otherwise return original"""
    if not word:
        return word
    
    # Check direct match first
    if word in PLURAL_MAP:
        return PLURAL_MAP[word]
    
    # Check case-insensitive match
    for singular, plural in PLURAL_MAP.items():
        if word.lower() == singular.lower():
            return plural
    
    # If not found, return original word
    return word

def run_generator(
    keywords_parent, keywords_child, new_apps, schedule_suffixes,
    delivery_manager, global_prod,
    rsp_duration, rsl_duration,
    sr_or_im, require_corp, require_recp, delivering_tag,
    support_group, managed_by_group, aliases_on, aliases_value,
    src_dir, out_dir,
    special_it=False, special_hr=False, special_medical=False, special_dak=False,
    use_custom_commitments=False, custom_commitments_str="", commitment_country=None,
    rsp_enabled=False, rsl_enabled=False,
    rsp_schedule="", rsl_schedule="",
    rsp_priority="", rsl_priority="",
    rsp_time="", rsl_time="",
    require_corp_it=False, require_corp_dedicated=False,
    use_new_parent=False, new_parent_offering="", new_parent="",
    keywords_excluded="",
    use_lvl2=False, service_type_lvl2="",
    support_groups_per_country=None, managed_by_groups_per_country=None,
    schedule_settings_per_country=None,
    use_custom_depend_on=False, custom_depend_on_value="",
    aliases_per_country=None,
    selected_languages=None,
    business_criticality="",
    approval_required=False,
    approval_required_value="empty",
    approval_groups_per_app=None,  # Add this parameter
    change_subscribed_location=False,
    custom_subscribed_location="Global",
    use_pluralization=True):  # Add this parameter with default True
    """
    Main generator function.
    """
    # Define helper function for cleaning approval values
    def clean_approval_value(val):
        if pd.isna(val) or str(val).strip() in ['', 'nan', 'NaN', 'None', 'none', 'NULL', 'null', '<NA>']:
            return 'false'
        val_str = str(val).strip().lower()  # Convert to lowercase here
        if val_str in ['true', 'false']:
            return val_str  # Already lowercase
        elif val_str in ['yes', 'y', '1']:
            return 'true'
        elif val_str in ['no', 'n', '0']:
            return 'false'
        else:
            # Keep custom values as-is but convert to lowercase
            return str(val).strip().lower()

    # Initialize per-country support groups dictionaries if not provided
    if support_groups_per_country is None:
        support_groups_per_country = {}
    if managed_by_groups_per_country is None:
        managed_by_groups_per_country = {}
    if schedule_settings_per_country is None:
        schedule_settings_per_country = {}
    if aliases_per_country is None:
        aliases_per_country = {}
    if selected_languages is None:
        selected_languages = []

    sheets, seen = {}, set()
    sheets_data = {}  # Store rows as lists for batch concatenation
    existing_offerings = set()  # Track existing offerings to detect duplicates
    original_ldap_data = {}  # Store LDAP data from original files
    excel_cache = {}  # Cache for Excel file reads
    missing_schedule_info = {}  # Track rows with missing schedules for red formatting
    column_order_cache = {}  # Store original column order from files

    def parse_keywords(keyword_string):
        """Parse keywords - returns (keywords_list, use_and_logic)"""
        if not keyword_string.strip():
            return [], False
        
        # Check if it contains commas (AND logic)
        if ',' in keyword_string:
            keywords = [k.strip() for k in keyword_string.split(',') if k.strip()]
            return keywords, True
        else:
            # Line separated (OR logic)
            keywords = [k.strip() for k in keyword_string.split('\n') if k.strip()]
            return keywords, False

    def row_keywords_ok(row):
        # Parse parent keywords
        parent_keywords, parent_use_and = parse_keywords(keywords_parent)
        
        # Check parent offering
        if parent_keywords:
            p = str(row["Parent Offering"]).lower()
            # Remove extra spaces and normalize
            p = ' '.join(p.split())
            
            if parent_use_and:
                # AND logic - all keywords must match (case-insensitive substring search)
                for k in parent_keywords:
                    if k.lower() not in p:
                        return False
            else:
                # OR logic - any keyword must match (case-insensitive substring search)
                found = False
                for k in parent_keywords:
                    if k.lower() in p:
                        found = True
                        break
                if not found:
                    return False
        
        # Parse child keywords
        child_keywords, child_use_and = parse_keywords(keywords_child)
        
        # Check child name
        if child_keywords:
            n = str(row["Name (Child Service Offering lvl 1)"]).lower()
            # Remove extra spaces and normalize
            n = ' '.join(n.split())
            
            if child_use_and:
                # AND logic - all keywords must match (case-insensitive substring search)
                for k in child_keywords:
                    if k.lower() not in n:
                        return False
            else:
                # OR logic - any keyword must match (case-insensitive substring search)
                found = False
                for k in child_keywords:
                    if k.lower() in n:
                        found = True
                        break
                if not found:
                    return False
        
        return True

    def row_excluded_keywords_ok(row):
        """Check if row should be excluded based on excluded keywords"""
        # Parse excluded keywords
        excluded_keywords, excluded_use_and = parse_keywords(keywords_excluded)
        
        if not excluded_keywords:
            return True  # No excluded keywords, so row is OK
        
        # Check both parent offering and child name for excluded keywords
        p = str(row["Parent Offering"]).lower()
        n = str(row["Name (Child Service Offering lvl 1)"]).lower()
        
        if excluded_use_and:
            # AND logic - if ALL excluded keywords are found, exclude the row
            parent_has_all = all(k.lower() in p for k in excluded_keywords)
            child_has_all = all(k.lower() in n for k in excluded_keywords)
            # Exclude if either parent or child has all excluded keywords
            if parent_has_all or child_has_all:
                return False
        else:
            # OR logic - if ANY excluded keyword is found, exclude the row
            parent_has_any = any(k.lower() in p for k in excluded_keywords)
            child_has_any = any(k.lower() in n for k in excluded_keywords)
            # Exclude if either parent or child has any excluded keyword
            if parent_has_any or child_has_any:
                return False
        
        return True  # Row is OK (not excluded)

    def lc_ok(row):
        return all(str(row[c]).strip().lower() not in discard_lc
                   for c in ("Phase", "Status", "Life Cycle Stage", "Life Cycle Status"))

    def name_prefix_ok(name):
        # Make prefix check case-insensitive and handle extra spaces
        name = name.strip()
        # Check if name starts with [SR or [IM (case-insensitive)
        return name.upper().startswith(f"[{sr_or_im.upper()} ") or name.upper().startswith(f"[{sr_or_im.upper()}\t")

    # Process apps - split on comma, newline, or semicolon
    all_apps = []
    for raw in new_apps:
        for app in re.split(r'[,\n;]+', str(raw)):
            app = app.strip()
            if app:
                all_apps.append(app)

    # If no apps provided, process without apps
    if not all_apps:
        all_apps = [None]  # Use None instead of empty string

    # Determine special department - this only affects naming, not filtering
    special_dept = None
    if special_it:
        special_dept = "IT"
    elif special_hr:
        special_dept = "HR"
    elif special_medical:
        special_dept = "Medical"
    elif special_dak:
        special_dept = "DAK"

    # First, collect all existing offerings, LDAP data, and column order from the source files
    for wb in src_dir.glob("ALL_Service_Offering_*.xlsx"):
        try:
            # Cache Excel file to avoid multiple reads
            if wb not in excel_cache:
                excel_cache[wb] = pd.ExcelFile(wb)
            excel_file = excel_cache[wb]
            
            # Check BOTH sheets for existing offerings and column order
            for sheet_name in ["Child SO lvl1", "Child SO lvl2"]:
                try:
                    if sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        
                        # Store column order
                        country = wb.stem.split("_")[-1].upper()
                        column_key = f"{country}_{sheet_name}"
                        if column_key not in column_order_cache:
                            column_order_cache[column_key] = list(df.columns)
                        
                        if "Name (Child Service Offering lvl 1)" in df.columns:
                            # Clean and normalize the names before adding to set
                            existing_names = df["Name (Child Service Offering lvl 1)"].dropna().astype(str)
                            normalized_names = existing_names.apply(lambda x: ' '.join(str(x).split()))
                            existing_offerings.update(normalized_names)
                        
                        # Collect LDAP data for DE
                        if wb.stem.endswith("_DE") and "Support group" in df.columns:
                            # Look for LDAP columns
                            ldap_cols = [col for col in df.columns if "LDAP" in col.upper() or "Ldap" in col or "ldap" in col]
                            if ldap_cols and "Support group" in df.columns:
                                for idx, row in df.iterrows():
                                    sg = str(row.get("Support group", "")).strip()
                                    if sg and sg not in ["nan", "NaN", ""]:
                                        # Store all LDAP values for this support group
                                        ldap_values = {}
                                        for ldap_col in ldap_cols:
                                            ldap_val = str(row.get(ldap_col, "")).strip()
                                            if ldap_val and ldap_val not in ["nan", "NaN", "", "None", "none"]:
                                                ldap_values[ldap_col] = ldap_val
                                        if ldap_values:
                                            original_ldap_data[sg] = ldap_values
                except Exception:
                    # Skip if sheet doesn't exist
                    continue
        except Exception:
            # Skip if there's an error reading the file
            continue

    # Process the files
    total_files = len(list(src_dir.glob("ALL_Service_Offering_*.xlsx")))
    processed_files = 0
    
    for wb in src_dir.glob("ALL_Service_Offering_*.xlsx"):
        processed_files += 1
        country = wb.stem.split("_")[-1].upper()
        print(f"Processing file {processed_files}/{total_files}: {wb.name}")
        
        # Process BOTH lvl1 and lvl2 sheets when use_lvl2 is True
        # Process only lvl1 when use_lvl2 is False
        levels_to_process = [1, 2] if use_lvl2 else [1]
        
        for current_level in levels_to_process:
            sheet_name = f"Child SO lvl{current_level}"
            is_lvl2 = (current_level == 2)
                
            try:
                # IF USING NEW PARENT, CREATE SYNTHETIC ROW
                if use_new_parent:
                    # Split the parent offerings and parents into separate lines
                    parent_offerings_list = [line.strip() for line in new_parent_offering.split('\n') if line.strip()]
                    parents_list = [line.strip() for line in new_parent.split('\n') if line.strip()]
                    
                    # Create multiple synthetic rows - one for each pair
                    synthetic_rows = []
                    for i in range(max(len(parent_offerings_list), len(parents_list))):
                        # Get the offering and parent for this index, or use the last available one
                        offering = parent_offerings_list[min(i, len(parent_offerings_list) - 1)] if parent_offerings_list else ""
                        parent = parents_list[min(i, len(parents_list) - 1)] if parents_list else ""
                        
                        # Create individual row with single values (not multi-line)
                        new_row = create_new_parent_row(offering, parent, country, business_criticality, approval_required, approval_required_value, change_subscribed_location, custom_subscribed_location)
                        synthetic_rows.append(new_row)
                    
                    # Create DataFrame from all synthetic rows
                    base_pool = pd.DataFrame(synthetic_rows)
                    
                    # Initialize schedule checking variables
                    all_country_names_for_schedules = pd.Series([], dtype=str)
                    corp_names_for_schedules = pd.Series([], dtype=str)
                    non_corp_names_for_schedules = pd.Series([], dtype=str)
                else:
                    # ORIGINAL LOGIC - read from cached Excel file
                    if wb not in excel_cache:
                        excel_cache[wb] = pd.ExcelFile(wb)
                    excel_file = excel_cache[wb]
                    
                    if sheet_name not in excel_file.sheet_names:
                        continue  # Skip if sheet doesn't exist
                        
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    
                    # Store the original column order for this sheet
                    column_key = f"{country}_{sheet_name}"
                    if column_key not in column_order_cache:
                        column_order_cache[column_key] = list(df.columns)
                    
                    # Add missing columns as empty, but maintain order
                    original_columns = list(df.columns)
                    
                    # DEBUG: Print all columns to see what we have
                    print(f"🔍 **COLUMNS IN {wb.name} - {sheet_name}**:")
                    for i, col in enumerate(df.columns):
                        print(f"  {i+1:2d}. '{col}'")
                        if "alias" in col.lower() or "u_label" in col.lower():
                            print(f"      ⭐ **ALIAS COLUMN FOUND!**")
                    
                    for col in need_cols:
                        if col not in df.columns:
                            df[col] = ""

                    # Remove the "Ensure Visibility group column exists" section
                    # if "Visibility group" not in df.columns:
                    #     df["Visibility group"] = ""

                    # Add LDAP columns if they don't exist
                    ldap_cols = [col for col in df.columns if "LDAP" in col.upper() or "Ldap" in col or "ldap" in col]
                    if not ldap_cols and country == "DE":
                        # Add a default LDAP column for DE
                        df["LDAP"] = ""
                        ldap_cols = ["LDAP"]

                    # Initialize schedule checking variables
                    all_country_names_for_schedules = pd.Series([], dtype=str)
                    corp_names_for_schedules = pd.Series([], dtype=str)
                    non_corp_names_for_schedules = pd.Series([], dtype=str)

                    # Debug: Check if we have the required columns
                    if "Parent Offering" in df.columns and "Name (Child Service Offering lvl 1)" in df.columns:
                        # Store ALL names from the country file for schedule checking (before any filtering)
                        all_country_names_for_schedules = df["Name (Child Service Offering lvl 1)"].astype(str).str.upper()
                        corp_names_for_schedules = all_country_names_for_schedules[all_country_names_for_schedules.str.contains('CORP|DEDICATED|RECP', na=False)]
                        non_corp_names_for_schedules = all_country_names_for_schedules[~all_country_names_for_schedules.str.contains('CORP|DEDICATED|RECP', na=False)]
                        
                    # Apply pre-filtering with vectorized operations where possible
                    if keywords_parent.strip():
                        # Fast pre-filter using vectorized string operations
                        parent_col = df["Parent Offering"].astype(str).str.lower()
                        parent_keywords, parent_use_and = parse_keywords(keywords_parent)
                        if parent_keywords:
                            if parent_use_and:
                                # AND logic - all keywords must be present
                                parent_mask = parent_col.str.contains('|'.join([re.escape(k.lower()) for k in parent_keywords]), na=False)
                                # Further filter with apply for exact AND logic
                                parent_mask = parent_mask & df.apply(row_keywords_ok, axis=1)
                            else:
                                # OR logic - any keyword must be present
                                parent_mask = parent_col.str.contains('|'.join([re.escape(k.lower()) for k in parent_keywords]), na=False)
                        else:
                            parent_mask = pd.Series([True] * len(df), index=df.index)
                        df = df[parent_mask]  # Reduce dataset size early
                    
                    if df.empty:
                        continue
                        
                    # Debug: Check how many rows match keywords before other filters
                    print(f"Rows after parent keyword pre-filter: {len(df)}")

                    # For Lvl2, different filtering logic
                    if is_lvl2:
                        # For Lvl2, we don't check for SR/IM prefix and handle commitments differently
                        mask = (df.apply(row_keywords_ok, axis=1)
                                & df.apply(row_excluded_keywords_ok, axis=1)
                                & df.apply(lc_ok, axis=1))
                        # Don't filter out entries with empty Service Commitments for Lvl2
                    else:
                        mask = (df.apply(row_keywords_ok, axis=1)
                                & df.apply(row_excluded_keywords_ok, axis=1)
                                & df["Name (Child Service Offering lvl 1)"].astype(str).apply(name_prefix_ok)
                                & df.apply(lc_ok, axis=1)
                                & (df["Service Commitments"].astype(str).str.strip().replace({"nan": ""}) != "-"))

                    base_pool = df.loc[mask]
                    print(f"Final matching rows: {len(base_pool)}")
                
                if base_pool.empty:
                    continue
                
                # Process ALL matching rows, not just the first one
                total_base_rows = len(base_pool)
                print(f"Processing {total_base_rows} base rows...")
                start_time = time.time()
                
                for row_idx, (idx, base_row) in enumerate(base_pool.iterrows()):
                    if row_idx % 10 == 0 and row_idx > 0:
                        elapsed = time.time() - start_time
                        avg_time = elapsed / row_idx
                        remaining = (total_base_rows - row_idx) * avg_time
                        print(f"  Processed {row_idx}/{total_base_rows} base rows... ETA: {remaining:.1f}s")
                        
                    base_row_df = base_row.to_frame().T.copy()
                    tag_hs, tag_ds = f"HS {country}", f"DS {country}"

                    # Define receivers ONLY ONCE based on country and context
                    if country == "PL":
                        receivers = ["HS PL", "DS PL"]
                    elif country == "CY":
                        receivers = ["DS CY"]  # CY now has only DS
                    elif country == "DE":
                        receivers = ["HS DE", "DS DE"]
                    elif country == "UA":
                        receivers = ["DS UA"]  # Only DS for UA
                    elif country == "MD":
                        receivers = ["DS MD"]  # Only DS for MD
                    elif country == "RO":
                        receivers = ["DS RO"]  # Only DS for RO
                    elif country == "TR":
                        receivers = ["DS TR"]  # Only DS for TR
                    else:
                        receivers = [f"HS {country}", f"DS {country}"]

                    parent_full = str(base_row["Parent Offering"])
                    
                    # Store original depend on value
                    original_depend_on = str(base_row.get("Service Offerings | Depend On (Application Service)", "")).strip()

                    for app in all_apps:
                        # DO NOT RECALCULATE receivers here! Use the ones defined above
                        for recv in receivers:
                            # Get country-specific schedule suffixes
                            country_schedule_suffixes = get_schedule_suffixes_for_country(
                                country, recv, schedule_settings_per_country, schedule_suffixes
                            )
                            
                            for schedule_suffix in country_schedule_suffixes:
                                # Check if schedule exists in the source data
                                missing_schedule = False
                                
                                # Prepare search patterns based on whether it's CORP or not
                                is_corp_type = (require_corp or require_recp or require_corp_it or require_corp_dedicated)
                                
                                # Only check schedules if we have data to check against
                                if len(all_country_names_for_schedules) > 0:
                                    # Normalize schedule for comparison
                                    schedule_pattern = schedule_suffix.strip()
                                    
                                    # Check based on CORP type
                                    if is_corp_type:
                                        # For CORP types, check only in CORP names
                                        if len(corp_names_for_schedules) > 0:
                                            # Look for exact schedule match in CORP offerings
                                            schedule_found = any(schedule_pattern in name for name in corp_names_for_schedules)
                                            if not schedule_found:
                                                missing_schedule = True
                                        else:
                                            # No CORP offerings exist, so schedule is missing
                                            missing_schedule = True
                                    else:
                                        # For non-CORP, check only in non-CORP names
                                        if len(non_corp_names_for_schedules) > 0:
                                            # Look for exact schedule match in non-CORP offerings
                                            schedule_found = any(schedule_pattern in name for name in non_corp_names_for_schedules)
                                            if not schedule_found:
                                                missing_schedule = True
                                        else:
                                            # No non-CORP offerings exist, so schedule is missing
                                            missing_schedule = True
                                
                                # For DE, find the matching row (DS DE or HS DE) in the original data
                                if country == "DE" and not use_new_parent:
                                    # Always attempt to pick matching row but do not skip if none found
                                    recv_mask = base_pool["Name (Child Service Offering lvl 1)"].str.contains(
                                        rf"\b{re.escape(recv)}\b", case=False
                                    )
                                    if recv_mask.any():
                                        # Use the first matching row as base
                                        base_row = base_pool[recv_mask].iloc[0]
                                        base_row_df = base_row.to_frame().T.copy()
                                        original_depend_on = str(base_row.get("Service Offerings | Depend On (Application Service)", "")).strip()
                        
                                # Build name based on type
                                if is_lvl2:
                                    new_name = build_lvl2_name(
                                        parent_full, sr_or_im, app, schedule_suffix, service_type_lvl2
                                    )
                                elif require_corp:
                                    new_name = build_corp_name(
                                        parent_full, sr_or_im, app, schedule_suffix, recv, delivering_tag
                                    )
                                elif require_recp:
                                    new_name = build_recp_name(
                                        parent_full, sr_or_im, app, schedule_suffix, recv, delivering_tag
                                    )
                                elif require_corp_it:
                                    new_name = build_corp_it_name(
                                        parent_full, sr_or_im, app, schedule_suffix, recv, delivering_tag
                                    )
                                elif require_corp_dedicated:
                                    new_name = build_corp_dedicated_name(
                                        parent_full, sr_or_im, app, schedule_suffix, recv, delivering_tag
                                    )
                                else:
                                    # For standard names, handle DE split naming
                                    if country == "DE" and recv:
                                        # Extract parent content to build proper name with DS/HS
                                        parent_content = extract_parent_info(parent_full)
                                        catalog_name = extract_catalog_name(parent_full)
                                        
                                        # Replace parent division with receiver division
                                        parts = parent_content.split()
                                        new_parts = [sr_or_im]
                                        
                                        # Add receiver division (DS or HS from recv)
                                        recv_division = recv.split()[0]  # Extract DS or HS from "DS DE" or "HS DE"
                                        new_parts.append(recv_division)
                                        
                                        # Add country and other parts
                                        for part in parts:
                                            if part in ["HS", "DS"]:
                                                continue  # Skip original division
                                            elif len(part) == 2 and part.isupper() and part not in ["IT", "HR"]:
                                                new_parts.append(part)  # Add country
                                            elif part in ["IT", "HR", "Medical", "Business Services"] or (part not in ["HS", "DS"] and not (len(part) == 2 and part.isupper())):
                                                new_parts.append(part)  # Add dept or other parts
                                        
                                        # Build new parent offering with updated division
                                        new_parent_offering_str = f"[Parent {' '.join(new_parts)}] {catalog_name}"
                                        new_name = build_standard_name(
                                            new_parent_offering_str, sr_or_im, app, schedule_suffix, special_dept, recv
                                        )
                                    else:
                                        new_name = build_standard_name(
                                            parent_full, sr_or_im, app, schedule_suffix, special_dept, recv
                                        )
                                
                                # Normalize the name for comparison (remove extra spaces)
                                new_name_normalized = ' '.join(new_name.split())
                                
                                # Check against existing offerings in source files
                                if new_name_normalized in existing_offerings:
                                    raise ValueError(f"Sorry, it would be a duplicate - we already have this offering in the system: {new_name}")
                                
                                # Determine division for PL support groups
                                division = None
                                if country == "PL":
                                    # Try to determine division from the new_name or original data
                                    if "HS PL" in new_name or "HS PL" in str(base_row.get("Name (Child Service Offering lvl 1)", "")):
                                        division = "HS"
                                    elif "DS PL" in new_name or "DS PL" in str(base_row.get("Name (Child Service Offering lvl 1)", "")):
                                        division = "DS"
                                    else:
                                        # Try to determine from parent offering
                                        parent_content = extract_parent_info(parent_full)
                                        if "HS" in parent_content.split():
                                            division = "HS"
                                        elif "DS" in parent_content.split():
                                            division = "DS"
                                        else:
                                            # Default to HS if cannot determine
                                            division = "HS"
                                
                                # Get support groups list
                                if country == "PL":
                                    # For PL, directly use the receiver-specific support group
                                    # The receiver is the key (e.g., "HS PL" or "DS PL")
                                    key = recv  # recv is already correctly set to "HS PL" or "DS PL"
                                    country_supports = support_groups_per_country.get(key, "")
                                    country_managed = managed_by_groups_per_country.get(key, "")
                                    
                                    # For PL, we expect only one support group per receiver
                                    if country_supports:
                                        sg = str(country_supports).strip()
                                        mg = str(country_managed or sg).strip()
                                        support_groups_list = [(sg, mg)]
                                    else:
                                        # Fallback to empty if no support group configured for this receiver
                                        support_groups_list = [("", "")]
                                else:
                                    # For other countries, use the existing logic
                                    support_groups_list = get_support_groups_list_for_country(
                                        country, support_group, support_groups_per_country, 
                                        managed_by_groups_per_country, division
                                    )
                                
                                # For DE, limit groups to those matching the current receiver if any, else keep all
                                if country == "DE" and recv:
                                    prefix = recv
                                    matching = [(sg, mg) for sg, mg in support_groups_list
                                                if sg.strip().startswith(prefix)]
                                    if matching:
                                        support_groups_list = matching
                                
                                # Create offerings for each support group combination
                                for support_group_for_country, managed_by_group_for_country in support_groups_list:
                                    # Skip duplicates based on name, receiver, app, schedule, support and managed groups
                                    key = (
                                        new_name_normalized,
                                        recv,
                                        app,
                                        schedule_suffix,
                                        support_group_for_country,
                                        managed_by_group_for_country
                                    )
                                    
                                    if key in seen:
                                        continue
                                    seen.add(key)
                                    row = base_row_df.copy()
                                    
                                    # Update the name
                                    row.loc[:, "Name (Child Service Offering lvl 1)"] = new_name
                                    row.loc[:, "Parent"] = new_parent if use_new_parent else ""
                                    row.loc[:, "Delivery Manager"] = delivery_manager
                                    
                                    # Apply business criticality if provided
                                    if business_criticality:
                                        row.loc[:, "Business Criticality"] = business_criticality
                                    # Otherwise, keep the original value from source file
                                    
                                    # Set Record view based on SR/IM selection
                                    if sr_or_im == "SR":
                                        row.loc[:, "Record view"] = "Request Item"
                                    elif sr_or_im == "IM":
                                        row.loc[:, "Record view"] = "Incident, Major Incident"
                                    
                                    # Set Approval required with conditional value
                                    if approval_required:
                                        row.loc[:, "Approval required"] = "true"  # Always use "true" when checkbox is ticked
                                        
                                        # Use per-app approval group if configured, otherwise use global value
                                        if approval_groups_per_app and app in approval_groups_per_app:
                                            app_approval_group = approval_groups_per_app[app].strip()
                                            # Keep empty if the user left it empty - don't force to "empty"
                                            row.loc[:, "Approval group"] = app_approval_group
                                        else:
                                            # If no per-app configuration exists, use global value (but not "PER_APP")
                                            if approval_required_value and approval_required_value != "PER_APP":
                                                row.loc[:, "Approval group"] = approval_required_value
                                            else:
                                                # Keep empty instead of forcing "empty"
                                                row.loc[:, "Approval group"] = ""
                                    else:
                                        row.loc[:, "Approval required"] = "false"
                                        row.loc[:, "Approval group"] = "empty"  # Use literal "empty" when not required
                                    
                                    # Set Subscribed by Location based on user choice
                                    if change_subscribed_location:
                                        row.loc[:, "Subscribed by Location"] = custom_subscribed_location
                                    else:
                                        row.loc[:, "Subscribed by Location"] = "Global"
                                    
                                    # Apply support group and managed by group
                                    row.loc[:, "Support group"] = support_group_for_country if support_group_for_country else ""
                                    row.loc[:, "Managed by Group"] = managed_by_group_for_country if managed_by_group_for_country else ""
                                    
                                    # Handle aliases
                                    exact_column_name = "Aliases (u_label) - ENG"
                                    if exact_column_name not in row.columns:
                                        row[exact_column_name] = ""
                                    if aliases_on and aliases_value == "USE_APP_NAMES":
                                        row.loc[:, exact_column_name] = app if app else ""
                                    else:
                                        # Kopiuj z oryginalnego pliku
                                        row.loc[:, exact_column_name] = base_row.get(exact_column_name, "")
                                    # Handle DE special cases
                                    if country == "DE":
                                        company, ldap = get_de_company_and_ldap(support_group_for_country, recv, base_row)
                                        row.loc[:, "Subscribed by Company"] = company
                                        
                                        # Handle LDAP columns
                                        ldap_cols = [col for col in row.columns if "LDAP" in col.upper() or "Ldap" in col or "ldap" in col]
                                        if ldap_cols:
                                            # Clear all LDAP columns first
                                            for ldap_col in ldap_cols:
                                                row.loc[:, ldap_col] = ""
                                            

                                    # Handle Subscribed by Company based on type and mode
                                    if use_new_parent:
                                        # NEW PARENT MODE - special logic
                                        if require_corp or require_recp or require_corp_it or require_corp_dedicated:
                                            # For CORP offerings, extract what comes after CORP
                                            # Example: [SR DS CY CORP HS DE Dedicated Services] -> "HS DE"
                                            match = re.search(r'\[.*?CORP\s+([A-Z]{2}\s+[A-Z]{2})', new_name)
                                            if match:
                                                row.loc[:, "Subscribed by Company"] = match.group(1)
                                            else:
                                                # Fallback to receiver if pattern not found
                                                row.loc[:, "Subscribed by Company"] = recv
                                        else:
                                            # For non-CORP in new parent mode, use receiver (e.g., "HS PL", "DS PL")
                                            row.loc[:, "Subscribed by Company"] = recv
                                    elif country == "DE":
                                        # Existing DE logic for Germany
                                        company, _ = get_de_company_and_ldap(support_group_for_country, recv, base_row)

                                        row.loc[:, "Subscribed by Company"] = company
                                    elif require_corp or require_recp or require_corp_it or require_corp_dedicated:
                                        # For CORP offerings in normal mode, clear the field
                                        row.loc[:, "Subscribed by Company"] = ""
                                    # For standard offerings, keep original value from source file
                                    
                                    
                                    orig_comm = str(row.iloc[0]["Service Commitments"]).strip()
                                    
                                    # If schedule is missing, use original commitments with user schedule
                                    if missing_schedule:
                                        if not orig_comm or orig_comm in ["-", "nan", "NaN", "", None]:
                                            # If original commitments are empty, create new ones with user schedule
                                            row.loc[:, "Service Commitments"] = commit_block(country, schedule_suffix, rsp_duration, rsl_duration, sr_or_im)
                                        else:
                                            # Update original commitments with user schedule
                                            row.loc[:, "Service Commitments"] = update_commitments(orig_comm, schedule_suffix, rsp_duration, rsl_duration, sr_or_im, country)
                                    # For Lvl2, keep empty commitments empty
                                    elif is_lvl2 and (not orig_comm or orig_comm in ["-", "nan", "NaN", "", None]):
                                        row.loc[:, "Service Commitments"] = ""
                                    else:
                                        # Handle custom commitments - use both approaches
                                        if use_custom_commitments and custom_commitments_str:
                                            # Use the direct string if provided
                                            row.loc[:, "Service Commitments"] = custom_commitments_str
                                        elif use_custom_commitments and commitment_country:
                                            # Use custom_commit_block function if country provided
                                            row.loc[:, "Service Commitments"] = custom_commit_block(
                                                commitment_country,

                                                sr_or_im, rsp_enabled, rsl_enabled,
                                                rsp_schedule, rsl_schedule, rsp_priority, rsl_priority,
                                                rsp_time, rsl_time
                                            )
                                        else:
                                            # Use existing logic
                                            if not orig_comm or orig_comm == "-":
                                                row.loc[:, "Service Commitments"] = commit_block(country, schedule_suffix, rsp_duration, rsl_duration, sr_or_im)
                                            else:
                                                row.loc[:, "Service Commitments"] = update_commitments(orig_comm, schedule_suffix, rsp_duration, rsl_duration, sr_or_im, country)
                                    
                                    # Special handling for IT with UA/MD/RO/TR - always use DS
                                    if (special_dept == "IT" or require_corp_it) and country in ["UA", "MD", "RO", "TR"]:
                                        depend_tag = f"DS {country} Prod"
                                    elif global_prod:
                                        depend_tag = "Global Prod"
                                    else:
                                        if country == "PL":
                                            # Regex-based PL Prod determination (case-insensitive)
                                            if re.search(r'\bHS\s+PL\b', new_name, re.IGNORECASE):
                                                depend_tag = "HS PL Prod"
                                            if re.search(r'\bHS\s+PL\b', new_name, re.IGNORECASE):
                                                depend_tag = "HS PL Prod"
                                            elif re.search(r'\bDS\s+PL\b', new_name, re.IGNORECASE):
                                                depend_tag = "DS PL Prod"
                                            else:
                                                depend_tag = "DS PL Prod"  # safe default
                                        elif recv:
                                            depend_tag = f"{recv} Prod"
                                        else:
                                            depend_tag = f"{delivering_tag} Prod" if (require_corp or require_recp or require_corp_it or require_corp_dedicated) else f"{tag_hs} Prod"
                                    
                                    # Always update Service Offerings | Depend On based on computed depend_tag and app
                                    if use_custom_depend_on and custom_depend_on_value:
                                        # Build the value using prefix + app name
                                        if app:
                                            # Apply pluralization to app name if enabled
                                            app_to_use = get_plural_form(app) if use_pluralization else app
                                            # Check if Global Prod is enabled
                                            if global_prod:
                                                # Replace the closing ] with Prod]
                                                prefix_with_prod = custom_depend_on_value.replace(']', ' Prod]')
                                                row.loc[:, "Service Offerings | Depend On (Application Service)"] = f"{prefix_with_prod} {app_to_use}"
                                            else:
                                                row.loc[:, "Service Offerings | Depend On (Application Service)"] = f"{custom_depend_on_value} {app_to_use}"
                                        else:
                                            # If no app, use just the prefix
                                            if global_prod:
                                                # Replace the closing ] with Prod]
                                                prefix_with_prod = custom_depend_on_value.replace(']', ' Prod]')
                                                row.loc[:, "Service Offerings | Depend On (Application Service)"] = prefix_with_prod
                                            else:
                                                row.loc[:, "Service Offerings | Depend On (Application Service)"] = custom_depend_on_value
                                    else:
                                        # If custom depend on is not enabled, leave the column empty regardless of app
                                        row.loc[:, "Service Offerings | Depend On (Application Service)"] = ""
                                

                                    
                                    # Create sheet key with level distinction
                                    sheet_key = f"{country} lvl{current_level}"
                                    
                                    
                                    # Get the column order for this sheet
                                    column_key = f"{country}_{sheet_name}"
                                    
                                    # Accumulate rows in lists for batch concatenation
                                    sheets_data.setdefault(sheet_key, [])
                                    if isinstance(row, pd.DataFrame):
                                        row_dict = row.iloc[0].to_dict()
                                    else:
                                        row_dict = row.to_dict()
                                    
                                    # Add missing schedule flag to the row dictionary
                                    if missing_schedule:
                                        row_dict["_missing_schedule"] = True
                                   
                                    else:
                                        row_dict["_missing_schedule"] = False
                                    
                                    # Store the column order key for this sheet
                                    row_dict["_column_order_key"] = column_key
                                    
                                    sheets_data[sheet_key].append(row_dict)
                            
            except Exception as e:
                # Skip if sheet doesn't exist or other error
                if "Worksheet" not in str(e):  # Only skip worksheet not found errors silently
                    print(f"Error processing {sheet_name} in {wb}: {e}")
                continue

    print(f"Starting main processing with {len(sheets_data)} potential sheets...")
    total_combinations = 0
    for sheet_key, rows_list in sheets_data.items():
        total_combinations += len(rows_list)
    print(f"Total combinations to process: {total_combinations}")

    # Convert lists to DataFrames once - major performance improvement!
    print(f"Converting {len(sheets_data)} sheet lists to DataFrames...")
    
    # Create output file path
    outfile = out_dir / f"Generated_Service_Offerings_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Ensure output directory exists
    out_dir.mkdir(parents=True, exist_ok=True)

    # Check if we have any data to write
    if not sheets_data or all(not rows_list for rows_list in sheets_data.values()):
        print("❌ No matching offerings found with the specified keywords.")
        print("Please check your search criteria:")
        if keywords_parent.strip():
            print(f"  - Parent keywords: {keywords_parent}")
        if keywords_child.strip():
            print(f"  - Child keywords: {keywords_child}")
        if keywords_excluded.strip():
            print(f"  - Excluded keywords: {keywords_excluded}")
        print("Try using different or fewer keywords, or check if the source files contain matching offerings.")
        
        # Return None or raise an exception instead of trying to create empty Excel
        raise ValueError("No matching offerings found. Please adjust your search criteria.")

    # Write to Excel with special handling for empty values
    with pd.ExcelWriter(outfile, engine="openpyxl") as w:
        sheets = {}  # Store final DataFrames for later use
        
        for sheet_key, rows_list in sheets_data.items():
            if rows_list:
                print(f"  Processing {sheet_key}: {len(rows_list)} rows")
                df = pd.DataFrame(rows_list)
                
                # Get the column order key from the first row
                column_order_key = None
                if "_column_order_key" in df.columns and len(df) > 0:
                    column_order_key = df.iloc[0]["_column_order_key"]

                # Extract sheet_name from sheet_key for use below
                sheet_name = sheet_key

                # Track which rows have missing schedules BEFORE dropping the column
                missing_schedule_rows = []
                if "_missing_schedule" in df.columns:
                    missing_schedule_rows = df[df["_missing_schedule"] == True].index.tolist()
                    # Store this info for later use
                    missing_schedule_info[sheet_name] = missing_schedule_rows
                
                # Remove helper columns
                for col in ["_missing_schedule", "_column_order_key"]:
                    if col in df.columns:
                        df = df.drop(columns=[col])
                
                # Reorder columns to match original order if we have it
                if column_order_key and column_order_key in column_order_cache:
                    original_order = column_order_cache[column_order_key]
                    
                    # Get the original DataFrame to preserve missing column values
                    original_df = None
                    if not use_new_parent:
                        # Try to get the original data for this sheet
                        country = sheet_key.split()[0]
                        level = sheet_key.split()[1]  # e.g., "lvl1"
                        sheet_name_lookup = f"Child SO {level}"
                        
                        # Find the source file for this country
                        for wb_path in src_dir.glob(f"ALL_Service_Offering_*{country}.xlsx"):
                            try:
                                if wb_path in excel_cache:
                                    excel_file = excel_cache[wb_path]
                                    if sheet_name_lookup in excel_file.sheet_names:
                                        original_df = pd.read_excel(excel_file, sheet_name=sheet_name_lookup)
                                        break
                            except Exception:
                                continue
                    
                    # Add missing columns from original, preserving their values - SAFER VERSION
                    for col in original_order:
                        if col not in df.columns:
                                df[col] = ''
                    
                    # Reorder columns to match original order, excluding Number column
                    ordered_cols = []
                    for col in original_order:
                        if col in df.columns and col != "Number":
                            ordered_cols.append(col)
                    
                    # Add any new columns that weren't in original
                    new_cols = [col for col in df.columns if col not in original_order]
                    
                    # Reorder DataFrame
                    df = df[ordered_cols + new_cols]
                
                # Extract country code from sheet_key (e.g., "PL lvl1" -> "PL")
                cc = sheet_key.split()[0]
                
                # Remove Visibility group column for PL only if it wasn't in the original
                # if cc == "PL" and "Visibility group" in df.columns:
                #     original_cols = column_order_cache.get(column_order_key, [])
                #     if "Visibility group" not in original_cols:
                #         df = df.drop(columns=["Visibility group"])
                
                # Clean data before writing to Excel - SAFER VERSION
                df_final = df.copy()
                
                # Clean cell values safely
                def _clean_cell(x):
                    """
                    Normalise cell values before saving to Excel.
                    Turns NaNs/None/NULL-like tokens into empty strings.
                    Leaves everything else untouched (but stripped).
                    """
                    if pd.isna(x):
                        return ''
                    s = str(x).strip()
                    if s.lower() in {'nan', 'none', 'null', '<na>', 'n/a'}:
                        return ''
                    return s

                # Use _clean_cell directly with applymap for DataFrame, apply for Series
                if isinstance(df_final, pd.DataFrame):
                    df_final = df_final.map(_clean_cell)
                elif isinstance(df_final, pd.Series):
                    df_final = df_final.apply(_clean_cell)

                # Keep the dedicated treatment for the two boolean-ish columns
                df_final["Approval required"] = df_final["Approval required"].apply(clean_approval_value)

                def _clean_approval_group(v):
                    v = _clean_cell(v)
                    return v  # Just return the cleaned value without forcing "empty"

                if "Approval group" in df_final.columns:
                    df_final["Approval group"] = df_final["Approval group"].apply(_clean_approval_group)
                
                # Store the final DataFrame for later formatting use
                sheets[sheet_key] = df_final
                
                # Write to Excel
                df_final.to_excel(w, sheet_name=sheet_key, index=False)
    
    # Apply formatting with red highlighting for missing schedules
    try:
        wb = load_workbook(outfile)
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        for ws in wb.worksheets:
            sheet_name = ws.title

            # Find the column index for "Name (Child Service Offering lvl 1)"
            name_col_idx = None
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == "Name (Child Service Offering lvl 1)":
                    name_col_idx = idx
                    break
            if name_col_idx is not None:
                name_col_letter = get_column_letter(name_col_idx)

                # Apply red formatting to each row with missing schedule
                for row_idx in missing_schedule_info.get(sheet_name, []):
                    # Excel rows are 1-based, and we need to account for header
                    excel_row_idx = row_idx + 2

                    # Check if row exists
                    if excel_row_idx <= ws.max_row:
                        cell = ws[f"{name_col_letter}{excel_row_idx}"]
                        cell.fill = red_fill

                        # Optionally, add a comment explaining why it's red (removed undefined variables)
                        # cell.comment = Comment(
                        #     "Schedule not found in source data for this offering",
                        #     "Service Offering Generator"
                        # )
            
            # Apply column widths and formatting
            for col_idx, col in enumerate(ws.columns, start=1):
                col_letter = get_column_letter(col_idx)
                
                # Calculate column width more safely
                max_length = 10  # minimum width
                for cell in col:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                          max_length = min(cell_length, 100)  # cap at 100 to prevent issues
                    except:
                        pass
                
                ws.column_dimensions[col_letter].width = max_length + 2
                
                # Apply text wrapping and clean cell values - SAFER VERSION
                for cell in col:
                    try:
                        if hasattr(cell, 'alignment'):
                            cell.alignment = Alignment(wrap_text=True)
                        
                        # Clean up cell values more safely
                        from openpyxl.cell.cell import MergedCell
                        if not isinstance(cell, MergedCell) and hasattr(cell, 'value') and cell.value is not None:
                            cell_val = str(cell.value).strip()
                            # Only clean up obviously invalid values
                            if cell_val.lower() in ['nan', 'none', 'null', '<na>', 'n/a'] or cell_val == '':
                                cell.value = None  # Use None instead of empty string for Excel
                            # Keep other values as-is to avoid corruption
                    except Exception as e:
                        # Don't print warnings for every cell - just continue
                        continue
        
        wb.save(outfile)
    except Exception as e:
        print(f"Warning: Error applying formatting: {e}")
        # If formatting fails, the file should still be valid without formatting
    
    # Clean up Excel cache to free memory
    for excel_file in excel_cache.values():
        try:
            if isinstance(excel_file, pd.ExcelFile):
                excel_file.close()
        except:
            pass

    print("Processing complete. Output saved to:", outfile)
    # Return the output file path
    return outfile